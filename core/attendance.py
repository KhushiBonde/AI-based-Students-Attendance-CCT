import os
import csv
import uuid
from datetime import datetime
import pandas as pd
import config

# CSV File Headers
STUDENT_HEADERS = [
    'roll_number', 'name', 'section', 'email', 'registered_at', 
    'consent_given', 'consent_timestamp', 'consent_ip', 
    'encoding_path', 'photo_count', 'active'
]

SESSION_HEADERS = [
    'session_id', 'subject', 'section', 'date', 'start_time', 
    'end_time', 'total_enrolled', 'total_present', 'csv_path'
]

ATTENDANCE_RECORD_HEADERS = [
    'session_id', 'roll_number', 'name', 'subject', 'section', 
    'date', 'timestamp', 'confidence_score', 'method', 'override_by'
]

def initialize_files():
    """Ensure master csv databases are initialized with headers."""
    if not os.path.exists(config.STUDENTS_CSV_PATH):
        with open(config.STUDENTS_CSV_PATH, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(STUDENT_HEADERS)
            
    sessions_csv_path = os.path.join(config.DATA_DIR, "sessions.csv")
    if not os.path.exists(sessions_csv_path):
        with open(sessions_csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(SESSION_HEADERS)

def write_student_to_master(roll_number, name, section, email, consent_ip, photo_count):
    """
    Register a student in students.csv.
    """
    initialize_files()
    
    # Check if student exists and update or add
    students = load_students()
    updated = False
    
    for i, s in enumerate(students):
        if s['roll_number'] == roll_number:
            students[i] = {
                'roll_number': roll_number,
                'name': name,
                'section': section,
                'email': email,
                'registered_at': s.get('registered_at', datetime.now().isoformat()),
                'consent_given': 'True',
                'consent_timestamp': datetime.now().isoformat(),
                'consent_ip': consent_ip,
                'encoding_path': f"db:{roll_number}",
                'photo_count': int(photo_count),
                'active': 'True'
            }
            updated = True
            break
            
    if not updated:
        students.append({
            'roll_number': roll_number,
            'name': name,
            'section': section,
            'email': email,
            'registered_at': datetime.now().isoformat(),
            'consent_given': 'True',
            'consent_timestamp': datetime.now().isoformat(),
            'consent_ip': consent_ip,
            'encoding_path': f"db:{roll_number}",
            'photo_count': int(photo_count),
            'active': 'True'
        })
        
    # Write back to CSV
    with open(config.STUDENTS_CSV_PATH, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=STUDENT_HEADERS)
        writer.writeheader()
        writer.writerows(students)

def load_students(active_only=True):
    """Load students list from students.csv."""
    initialize_files()
    students = []
    with open(config.STUDENTS_CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if active_only and row.get('active') != 'True':
                continue
            students.append(row)
    return students

def soft_delete_student(roll_number):
    """Soft delete student in the master registry."""
    students = load_students(active_only=False)
    found = False
    for s in students:
        if s['roll_number'] == roll_number:
            s['active'] = 'False'
            found = True
            
    if found:
        with open(config.STUDENTS_CSV_PATH, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=STUDENT_HEADERS)
            writer.writeheader()
            writer.writerows(students)
        return True
    return False


class SessionManager:
    """Manages active session, logs records to memory and CSV files."""
    _active_session = None

    @classmethod
    def start_session(cls, subject, section):
        if cls._active_session is not None:
            return False, "A session is already running."
            
        initialize_files()
        
        session_id = str(uuid.uuid4())
        date_str = datetime.now().strftime('%Y-%m-%d')
        start_time = datetime.now().isoformat()
        
        # Format session CSV filename
        filename = f"{date_str}_{subject.replace(' ', '_')}_{section.replace(' ', '_')}_{session_id[:8]}.csv"
        csv_path = os.path.join(config.ATTENDANCE_DIR, filename)
        
        # Initialize session CSV file
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(ATTENDANCE_RECORD_HEADERS)
            
        # Get count of enrolled students in this section
        all_students = load_students()
        enrolled_count = sum(1 for s in all_students if s['section'] == section)
        
        cls._active_session = {
            'session_id': session_id,
            'subject': subject,
            'section': section,
            'date': date_str,
            'start_time': start_time,
            'csv_path': csv_path,
            'total_enrolled': enrolled_count,
            'records': {}  # roll_number -> record dict
        }
        
        return True, cls._active_session

    @classmethod
    def get_active_session(cls):
        return cls._active_session

    @classmethod
    def mark_attendance(cls, roll_number, name, confidence_score, method='auto', override_by=''):
        """Mark student present with duplicate check and direct CSV write."""
        if cls._active_session is None:
            return False, "No active session."
            
        session = cls._active_session
        
        # Duplicate check
        if roll_number in session['records'] and method == 'auto':
            return False, "Already marked present."
            
        timestamp = datetime.now().isoformat()
        record = {
            'session_id': session['session_id'],
            'roll_number': roll_number,
            'name': name,
            'subject': session['subject'],
            'section': session['section'],
            'date': session['date'],
            'timestamp': timestamp,
            'confidence_score': float(confidence_score),
            'method': method,
            'override_by': override_by
        }
        
        session['records'][roll_number] = record
        
        # Write directly to session CSV (crash resilient)
        with open(session['csv_path'], 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=ATTENDANCE_RECORD_HEADERS)
            writer.writerow(record)
            
        return True, record

    @classmethod
    def mark_absent_override(cls, roll_number):
        """Remove a student from the active session records (manual override)."""
        if cls._active_session is None:
            return False, "No active session."
            
        session = cls._active_session
        if roll_number in session['records']:
            del session['records'][roll_number]
            
            # Rewrite the CSV file to reflect the change
            records_list = list(session['records'].values())
            with open(session['csv_path'], 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=ATTENDANCE_RECORD_HEADERS)
                writer.writeheader()
                writer.writerows(records_list)
            return True, "Student marked absent."
        return False, "Student was not marked present."

    @classmethod
    def stop_session(cls):
        """Finalize active session and save details in sessions.csv index."""
        if cls._active_session is None:
            return False, "No active session."
            
        session = cls._active_session
        end_time = datetime.now().isoformat()
        total_present = len(session['records'])
        
        # Save session to master session logs
        sessions_csv_path = os.path.join(config.DATA_DIR, "sessions.csv")
        session_entry = {
            'session_id': session['session_id'],
            'subject': session['subject'],
            'section': session['section'],
            'date': session['date'],
            'start_time': session['start_time'],
            'end_time': end_time,
            'total_enrolled': session['total_enrolled'],
            'total_present': total_present,
            'csv_path': session['csv_path']
        }
        
        with open(sessions_csv_path, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=SESSION_HEADERS)
            writer.writerow(session_entry)
            
        cls._active_session = None
        return True, session_entry

def get_session_history():
    """Get history of all finished sessions."""
    initialize_files()
    sessions = []
    sessions_csv_path = os.path.join(config.DATA_DIR, "sessions.csv")
    if os.path.exists(sessions_csv_path):
        with open(sessions_csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                sessions.append(row)
    # Return sorted by date/start_time descending
    return sorted(sessions, key=lambda x: x.get('start_time', ''), reverse=True)

def delete_session(session_id):
    """Delete a session's records from index and delete the csv file."""
    sessions_csv_path = os.path.join(config.DATA_DIR, "sessions.csv")
    if not os.path.exists(sessions_csv_path):
        return False
        
    sessions = []
    target_csv = None
    deleted = False
    
    with open(sessions_csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['session_id'] == session_id:
                target_csv = row['csv_path']
                deleted = True
            else:
                sessions.append(row)
                
    if deleted:
        # Save updated index
        with open(sessions_csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=SESSION_HEADERS)
            writer.writeheader()
            writer.writerows(sessions)
            
        # Delete the actual CSV file
        if target_csv and os.path.exists(target_csv):
            try:
                os.remove(target_csv)
            except Exception as e:
                print(f"Error deleting CSV file {target_csv}: {e}")
        return True
    return False

def generate_weekly_report_excel(output_path):
    """
    Generate Excel weekly report grouping attendance per student per subject,
    flagging students below 75% in red.
    """
    initialize_files()
    
    # Find all session files
    files = [os.path.join(config.ATTENDANCE_DIR, f) for f in os.listdir(config.ATTENDANCE_DIR) if f.endswith('.csv')]
    if not files:
        # Create an empty dataframe to save
        df_empty = pd.DataFrame(columns=['Roll Number', 'Name', 'Section', 'Subject', 'Sessions Attended', 'Total Sessions', 'Attendance %'])
        df_empty.to_excel(output_path, index=False)
        return False, "No attendance data found."
        
    # Read all csvs
    all_data = []
    for file in files:
        try:
            df = pd.read_csv(file)
            if not df.empty:
                all_data.append(df)
        except Exception:
            continue
            
    if not all_data:
        df_empty = pd.DataFrame(columns=['Roll Number', 'Name', 'Section', 'Subject', 'Sessions Attended', 'Total Sessions', 'Attendance %'])
        df_empty.to_excel(output_path, index=False)
        return False, "No records inside attendance files."
        
    df_att = pd.concat(all_data, ignore_index=True)
    df_att['roll_number'] = df_att['roll_number'].astype(str)
    df_att['timestamp'] = pd.to_datetime(df_att['timestamp'])
    
    # Calculate counts of sessions attended per student per subject
    attended = df_att.groupby(['roll_number', 'name', 'section', 'subject']).size().reset_index(name='attended')
    
    # Calculate total sessions held per subject/section from sessions.csv
    sessions_history = get_session_history()
    if not sessions_history:
        return False, "No session history recorded."
        
    df_sess = pd.DataFrame(sessions_history)
    session_totals = df_sess.groupby(['subject', 'section']).size().reset_index(name='total_sessions')
    
    # Join student list to get all registered students so we show 0% for those who didn't attend any
    students = load_students()
    df_students = pd.DataFrame(students)
    if df_students.empty:
        return False, "No students registered."
    df_students['roll_number'] = df_students['roll_number'].astype(str)
        
    # Generate all combinations of (student) x (subjects held for their section)
    # Find what subjects exist for each section
    section_subjects = df_sess[['section', 'subject']].drop_duplicates()
    
    # Merge students with their section's subjects
    base_df = pd.merge(df_students[['roll_number', 'name', 'section']], section_subjects, on='section')
    
    # Merge with attended sessions
    report_df = pd.merge(base_df, attended, on=['roll_number', 'name', 'section', 'subject'], how='left')
    report_df['attended'] = report_df['attended'].fillna(0).astype(int)
    
    # Merge with total sessions held
    report_df = pd.merge(report_df, session_totals, on=['subject', 'section'], how='left')
    report_df['total_sessions'] = report_df['total_sessions'].fillna(0).astype(int)
    
    # Calculate attendance %
    report_df['attendance_pct'] = report_df.apply(
        lambda r: round((r['attended'] / r['total_sessions'] * 100), 1) if r['total_sessions'] > 0 else 100.0,
        axis=1
    )
    
    # Rename columns for presentation
    report_df.columns = ['Roll Number', 'Name', 'Section', 'Subject', 'Sessions Attended', 'Total Sessions', 'Attendance %']
    
    # Save to Excel and format with openpyxl
    report_df.to_excel(output_path, index=False, sheet_name='Weekly Attendance')
    
    # Load workbook to apply red highlights for < 75%
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006', bold=True)
    
    # Find column index for Attendance %
    pct_col_idx = 7 # G is the 7th column
    
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=pct_col_idx)
        try:
            val = float(cell.value)
            if val < 75.0:
                cell.fill = red_fill
                cell.font = red_font
        except (ValueError, TypeError):
            continue
            
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_path)
    return True, "Excel report successfully created."
